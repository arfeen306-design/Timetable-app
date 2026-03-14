"""Excel export with formatted timetable sheets."""
from __future__ import annotations
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle,
)
from openpyxl.utils import get_column_letter

from repositories.timetable_repo import TimetableRepository
from repositories.class_repo import ClassRepository
from repositories.teacher_repo import TeacherRepository
from repositories.room_repo import RoomRepository
from repositories.school_repo import SchoolRepository
from repositories.lesson_repo import LessonRepository
from services.class_service import ClassService
from services.teacher_service import TeacherService
from utils.helpers import get_day_short, get_period_label, get_period_label_short, get_period_times, format_time_range, get_day_slot_sequence, color_hex_to_rgb, card_colors

if TYPE_CHECKING:
    from database.connection import DatabaseConnection


def export_excel(db: DatabaseConnection, path: str) -> None:
    """Export complete timetable workbook."""
    wb = Workbook()

    school_repo = SchoolRepository(db)
    school = school_repo.get()
    if not school:
        raise ValueError("No school settings found.")

    num_days = school.days_per_week
    num_periods_base = school.periods_per_day
    import json
    bell = {}
    if school.bell_schedule_json:
        try:
            raw = school.bell_schedule_json
            bell = json.loads(raw) if isinstance(raw, str) else raw
            if not isinstance(bell, dict):
                bell = {}
        except (TypeError, ValueError):
            pass
    zero_period = bool(bell.get("zero_period", False))
    num_periods = num_periods_base + (1 if zero_period else 0)

    slots_per_day = [
        get_day_slot_sequence(bell, d, num_periods, zero_period)
        for d in range(num_days)
    ]
    num_slots = max(len(s) for s in slots_per_day) if slots_per_day else num_periods
    vertical_labels = []
    for r in range(num_slots):
        if slots_per_day and r < len(slots_per_day[0]):
            slot = slots_per_day[0][r]
            if slot.get("type") == "period":
                lbl = get_period_label_short(slot["period_index"], zero_period)
                vertical_labels.append(f"{lbl}\n{format_time_range(slot.get('start', ''), slot.get('end', ''))}")
            else:
                vertical_labels.append(f"{slot.get('name', 'Break')}\n{format_time_range(slot.get('start', ''), slot.get('end', ''))}")
        else:
            vertical_labels.append("—")

    tt_repo = TimetableRepository(db)
    class_repo = ClassRepository(db)
    teacher_repo = TeacherRepository(db)
    room_repo = RoomRepository(db)
    lesson_repo = LessonRepository(db)

    # Common styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    title_font = Font(bold=True, size=14)

    # --- Class timetables --- (academic order: same as UI)
    classes = ClassService(db).get_all()
    for cls in classes:
        ws = wb.create_sheet(title=f"C-{cls.name}"[:31])
        entries = tt_repo.get_by_class(cls.id)

        # Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_slots + 1)
        ws.cell(1, 1, f"Timetable: {cls.name}").font = title_font

        # School name
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_slots + 1)
        ws.cell(2, 1, f"{school.name} - {school.academic_year}")

        # Headers: Day | slot1 | slot2 | ...
        row_start = 4
        ws.cell(row_start, 1, "Day").font = header_font
        ws.cell(row_start, 1).fill = header_fill
        ws.cell(row_start, 1).alignment = header_align
        ws.cell(row_start, 1).border = thin_border
        ws.column_dimensions["A"].width = 10
        for col in range(num_slots):
            c = col + 2
            cell = ws.cell(row_start, c, vertical_labels[col] if col < len(vertical_labels) else "—")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(c)].width = 14

        # Build grid: one row per day
        entry_map: dict[tuple[int, int], object] = {}
        for e in entries:
            entry_map[(e.day_index, e.period_index)] = e

        break_fill = PatternFill("solid", fgColor="E8E8E8")
        for day in range(num_days):
            row = row_start + 1 + day
            ws.cell(row, 1, get_day_short(day)).font = Font(bold=True)
            ws.cell(row, 1).alignment = cell_align
            ws.cell(row, 1).border = thin_border
            day_slots = slots_per_day[day] if day < len(slots_per_day) else []
            for col in range(num_slots):
                c = col + 2
                cell = ws.cell(row, c)
                cell.border = thin_border
                cell.alignment = cell_align
                if col >= len(day_slots):
                    continue
                slot = day_slots[col]
                if slot.get("type") == "period":
                    p = slot.get("period_index", 0)
                    entry = entry_map.get((day, p))
                    if entry:
                        text = entry.subject_code or entry.subject_name
                        if entry.teacher_name:
                            text += f"\n{entry.teacher_name}"
                        if entry.room_name:
                            text += f"\n{entry.room_name}"
                        cell.value = text
                        bg_hex, fg_hex = card_colors(entry.subject_color)
                        r, g, b = color_hex_to_rgb(bg_hex)
                        cell.fill = PatternFill("solid", fgColor=f"{r:02X}{g:02X}{b:02X}")
                        cell.font = Font(color=fg_hex.lstrip("#"), size=10)
                else:
                    cell.value = f"{slot.get('name', 'Break')}\n{format_time_range(slot.get('start', ''), slot.get('end', ''))}"
                    cell.fill = break_fill

    lessons = lesson_repo.get_all()

    # --- Teacher timetables --- (same order as UI)
    teachers = TeacherService(db).get_all()
    for teacher in teachers:
        ws = wb.create_sheet(title=f"T-{teacher.display_name}"[:31])
        entries = tt_repo.get_by_teacher(teacher.id)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_slots + 1)
        ws.cell(1, 1, f"Timetable: {teacher.full_name}").font = title_font

        row_start = 3
        ws.cell(row_start, 1, "Day").font = header_font
        ws.cell(row_start, 1).fill = header_fill
        ws.cell(row_start, 1).alignment = header_align
        ws.cell(row_start, 1).border = thin_border
        ws.column_dimensions["A"].width = 10
        for col in range(num_slots):
            c = col + 2
            cell = ws.cell(row_start, c, vertical_labels[col] if col < len(vertical_labels) else "—")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(c)].width = 14

        entry_map = {}
        for e in entries:
            entry_map[(e.day_index, e.period_index)] = e

        break_fill = PatternFill("solid", fgColor="E8E8E8")
        for day in range(num_days):
            row = row_start + 1 + day
            ws.cell(row, 1, get_day_short(day)).font = Font(bold=True)
            ws.cell(row, 1).alignment = cell_align
            ws.cell(row, 1).border = thin_border
            day_slots = slots_per_day[day] if day < len(slots_per_day) else []
            for col in range(num_slots):
                c = col + 2
                cell = ws.cell(row, c)
                cell.border = thin_border
                cell.alignment = cell_align
                if col >= len(day_slots):
                    continue
                slot = day_slots[col]
                if slot.get("type") == "period":
                    p = slot.get("period_index", 0)
                    entry = entry_map.get((day, p))
                    if entry:
                        text = f"{entry.subject_code or entry.subject_name}\n{entry.class_name}"
                        if entry.room_name:
                            text += f"\n{entry.room_name}"
                        cell.value = text
                        bg_hex, fg_hex = card_colors(entry.subject_color)
                        r, g, b = color_hex_to_rgb(bg_hex)
                        cell.fill = PatternFill("solid", fgColor=f"{r:02X}{g:02X}{b:02X}")
                        cell.font = Font(color=fg_hex.lstrip("#"), size=10)
                else:
                    cell.value = f"{slot.get('name', 'Break')}\n{format_time_range(slot.get('start', ''), slot.get('end', ''))}"
                    cell.fill = break_fill

        # Total weekly periods at bottom of teacher sheet (consistent with Teacher Load and PDF)
        total_periods = sum(l.periods_per_week for l in lessons if l.teacher_id == teacher.id)
        workload_row = row_start + num_days + 2
        ws.cell(workload_row, 1, "Total weekly periods:").font = Font(bold=True)
        ws.cell(workload_row, 2, total_periods)

    # --- Master Timetable (all classes vertically) ---
    ws_master = wb.create_sheet(title="Master Timetable")
    row = 1
    ws_master.cell(row, 1, f"{school.name} - {school.academic_year}").font = title_font
    row += 1
    ws_master.cell(row, 1, "Master Timetable — All Classes").font = Font(bold=True, size=12)
    row += 2

    for cls in classes:
        entries = tt_repo.get_by_class(cls.id)
        entry_map = {}
        for e in entries:
            entry_map[(e.day_index, e.period_index)] = e

        ws_master.cell(row, 1, f"Class: {cls.name}").font = Font(bold=True)
        row += 1
        # Header row: Day | slot1 | slot2 | ...
        ws_master.cell(row, 1, "Day").font = header_font
        ws_master.cell(row, 1).fill = header_fill
        ws_master.cell(row, 1).alignment = header_align
        for col in range(num_slots):
            c = col + 2
            cell = ws_master.cell(row, c, vertical_labels[col] if col < len(vertical_labels) else "—")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        row += 1
        break_fill_master = PatternFill("solid", fgColor="E8E8E8")
        for day in range(num_days):
            ws_master.cell(row, 1, get_day_short(day)).font = Font(bold=True)
            day_slots = slots_per_day[day] if day < len(slots_per_day) else []
            for col in range(num_slots):
                entry = None
                slot = day_slots[col] if col < len(day_slots) else None
                if slot and slot.get("type") == "period":
                    entry = entry_map.get((day, slot.get("period_index", 0)))
                cell = ws_master.cell(row, col + 2)
                cell.alignment = cell_align
                if entry:
                    cell.value = entry.subject_code or entry.subject_name
                    bg_hex, fg_hex = card_colors(entry.subject_color)
                    r, g, b = color_hex_to_rgb(bg_hex)
                    cell.fill = PatternFill("solid", fgColor=f"{r:02X}{g:02X}{b:02X}")
                    cell.font = Font(color=fg_hex.lstrip("#"), size=9)
                elif slot and slot.get("type") != "period":
                    cell.value = f"{slot.get('name', 'Break')} {format_time_range(slot.get('start', ''), slot.get('end', ''))}"
                    cell.fill = break_fill_master
            row += 1
        row += 1  # gap between classes

    # --- Teacher Load Summary --- (label consistent with teacher sheet and PDF)
    ws = wb.create_sheet(title="Teacher Load")
    ws.cell(1, 1, "Teacher Load Summary").font = title_font
    ws.cell(3, 1, "Teacher").font = Font(bold=True)
    ws.cell(3, 2, "Weekly periods").font = Font(bold=True)
    ws.cell(3, 3, "Max/Week").font = Font(bold=True)
    ws.cell(3, 4, "Max/Day").font = Font(bold=True)
    ws.column_dimensions["A"].width = 25

    for i, teacher in enumerate(teachers):
        row = 4 + i
        ws.cell(row, 1, teacher.full_name)
        total = sum(l.periods_per_week for l in lessons if l.teacher_id == teacher.id)
        ws.cell(row, 2, total)
        ws.cell(row, 3, teacher.max_periods_week)
        ws.cell(row, 4, teacher.max_periods_day)

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(path)
