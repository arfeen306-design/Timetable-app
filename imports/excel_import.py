"""Excel import for teachers, classes, and subjects with row-level validation."""
from __future__ import annotations
from dataclasses import dataclass, field
from typing import TYPE_CHECKING

from openpyxl import load_workbook

if TYPE_CHECKING:
    from database.connection import DatabaseConnection


@dataclass
class RowError:
    row: int  # 1-based Excel row
    message: str


@dataclass
class ImportResult:
    success_count: int = 0
    errors: list[RowError] = field(default_factory=list)
    total_rows: int = 0  # data rows (excluding header) considered

    @property
    def is_valid(self) -> bool:
        return len(self.errors) == 0

    @property
    def invalid_count(self) -> int:
        return len(self.errors)


def _cell(row: list, col_index: int):
    if col_index < len(row):
        v = row[col_index]
        return str(v).strip() if v is not None else ""
    return ""


def import_teachers_from_excel(db: "DatabaseConnection", path: str, *, dry_run: bool = False) -> ImportResult:
    """Import teachers from Excel. If dry_run=True, only validate and return counts/errors; no DB write."""
    result = ImportResult()
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        result.errors.append(RowError(1, "File must have a header row and at least one data row."))
        wb.close()
        return result

    data_rows = rows[1:]
    result.total_rows = len(data_rows)

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    first_i = next((i for i, h in enumerate(header) if "first" in h and "name" in h), 0)
    last_i = next((i for i, h in enumerate(header) if "last" in h and "name" in h), 1)
    code_i = next((i for i, h in enumerate(header) if "abbreviation" in h or "code" in h), 2)
    title_i = next((i for i, h in enumerate(header) if "title" in h), 3)
    max_day_i = next((i for i, h in enumerate(header) if "max" in h and "day" in h), -1)
    max_week_i = next((i for i, h in enumerate(header) if "max" in h and "week" in h), -1)

    for row_idx, row in enumerate(rows[1:], start=2):
        row = list(row) if row else []
        first = _cell(row, first_i)
        last = _cell(row, last_i)
        if not first and not last:
            continue
        if not first:
            result.errors.append(RowError(row_idx, "First Name is required."))
            continue
        code = _cell(row, code_i) or first[:3].upper()
        title = _cell(row, title_i) or "Mr."
        try:
            max_day = int(_cell(row, max_day_i) or 6)
            max_week = int(_cell(row, max_week_i) or 30)
        except ValueError:
            result.errors.append(RowError(row_idx, "Max Periods Day/Week must be numbers."))
            continue
        if max_day < 1 or max_week < 1:
            result.errors.append(RowError(row_idx, "Max periods must be at least 1."))
            continue
        existing = db.fetchone(
            "SELECT id FROM teacher WHERE LOWER(TRIM(first_name)) = LOWER(?) "
            "AND LOWER(TRIM(COALESCE(last_name, ''))) = LOWER(?) LIMIT 1",
            (first, last),
        )
        if existing:
            result.errors.append(
                RowError(row_idx, "A teacher with this first and last name already exists.")
            )
            continue
        if not dry_run:
            try:
                db.execute(
                    "INSERT INTO teacher (first_name, last_name, code, title, max_periods_day, max_periods_week) "
                    "VALUES (?,?,?,?,?,?)",
                    (first, last, code, title, max_day, max_week),
                )
                result.success_count += 1
            except Exception as e:
                result.errors.append(RowError(row_idx, str(e)))
        else:
            result.success_count += 1
    if not dry_run:
        db.commit()
    wb.close()
    return result


def import_classes_from_excel(db: "DatabaseConnection", path: str, *, dry_run: bool = False) -> ImportResult:
    """Import classes. If dry_run=True, only validate and return counts/errors; no DB write."""
    result = ImportResult()
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        result.errors.append(RowError(1, "File must have a header row and at least one data row."))
        wb.close()
        return result

    result.total_rows = len(rows) - 1

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    grade_i = next((i for i, h in enumerate(header) if "grade" in h), 0)
    section_i = next((i for i, h in enumerate(header) if "section" in h), 1)
    stream_i = next((i for i, h in enumerate(header) if "stream" in h), -1)
    name_i = next((i for i, h in enumerate(header) if "name" in h), 2)
    if name_i >= len(header):
        name_i = max(grade_i, section_i, stream_i if stream_i >= 0 else 0) + 1

    for row_idx, row in enumerate(rows[1:], start=2):
        row = list(row) if row else []
        grade = _cell(row, grade_i)
        section = _cell(row, section_i)
        stream = _cell(row, stream_i) if stream_i >= 0 else ""
        name = _cell(row, name_i)
        if not name:
            name = f"Grade {grade} {section}".strip() or f"Row {row_idx}"
        existing = db.fetchone(
            "SELECT id FROM school_class WHERE LOWER(TRIM(grade)) = LOWER(?) "
            "AND LOWER(TRIM(COALESCE(section, ''))) = LOWER(?) "
            "AND LOWER(TRIM(COALESCE(stream, ''))) = LOWER(?) LIMIT 1",
            (grade or "", section or "", stream or ""),
        )
        if existing:
            result.errors.append(
                RowError(row_idx, "A class with this grade, section and stream already exists.")
            )
            continue
        if not dry_run:
            try:
                code = (str(grade or "") + str(section or "")).replace(" ", "")[:10] or name[:6]
                db.execute(
                    "INSERT INTO school_class (grade, section, stream, name, code, strength) VALUES (?,?,?,?,?,?)",
                    (grade or "", section or "", stream or "", name, code, 30),
                )
                result.success_count += 1
            except Exception as e:
                result.errors.append(RowError(row_idx, str(e)))
        else:
            result.success_count += 1
    if not dry_run:
        db.commit()
    wb.close()
    return result


def import_subjects_from_excel(db: "DatabaseConnection", path: str, *, dry_run: bool = False) -> ImportResult:
    """Import subjects. If dry_run=True, only validate and return counts/errors; no DB write."""
    result = ImportResult()
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        result.errors.append(RowError(1, "File must have a header row and at least one data row."))
        wb.close()
        return result

    result.total_rows = len(rows) - 1

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    name_i = next((i for i, h in enumerate(header) if "name" in h), 0)
    code_i = next((i for i, h in enumerate(header) if "code" in h or "abbreviation" in h), 1)
    category_i = next((i for i, h in enumerate(header) if "category" in h), -1)
    max_day_i = next((i for i, h in enumerate(header) if "max" in h and "day" in h), -1)

    for row_idx, row in enumerate(rows[1:], start=2):
        row = list(row) if row else []
        name = _cell(row, name_i)
        if not name:
            continue
        code = _cell(row, code_i) or name[:3].upper()
        category = _cell(row, category_i) if category_i >= 0 else "Core"
        if category not in ("Core", "Elective", "Lab", "Activity", "Language", "Other"):
            category = "Core"
        try:
            max_per_day = int(_cell(row, max_day_i) or 2) if max_day_i >= 0 else 2
        except ValueError:
            result.errors.append(RowError(row_idx, "Max Per Day must be a number."))
            continue
        existing = db.fetchone(
            "SELECT id FROM subject WHERE LOWER(TRIM(name)) = LOWER(?) OR LOWER(TRIM(code)) = LOWER(?) LIMIT 1",
            (name, code),
        )
        if existing:
            result.errors.append(
                RowError(row_idx, "A subject with this name or code already exists.")
            )
            continue
        if not dry_run:
            try:
                db.execute(
                    "INSERT INTO subject (name, code, category, max_per_day) VALUES (?,?,?,?)",
                    (name, code, category, max_per_day),
                )
                result.success_count += 1
            except Exception as e:
                result.errors.append(RowError(row_idx, str(e)))
        else:
            result.success_count += 1
    if not dry_run:
        db.commit()
    wb.close()
    return result


def import_lessons_from_excel(db: "DatabaseConnection", path: str, *, dry_run: bool = False) -> ImportResult:
    """Import lesson mappings. If dry_run=True, only validate and return counts/errors; no DB write."""
    result = ImportResult()
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        result.errors.append(RowError(1, "File must have a header row and at least one data row."))
        wb.close()
        return result

    result.total_rows = len(rows) - 1

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    teacher_i = next((i for i, h in enumerate(header) if "teacher" in h), 0)
    subject_i = next((i for i, h in enumerate(header) if "subject" in h), 1)
    class_i = next((i for i, h in enumerate(header) if "class" in h), 2)
    periods_i = next((i for i, h in enumerate(header) if "period" in h and "week" in h), 3)
    duration_i = next((i for i, h in enumerate(header) if "duration" in h or "length" in h), -1)

    for row_idx, row in enumerate(rows[1:], start=2):
        row = list(row) if row else []
        teacher_val = _cell(row, teacher_i)
        subject_val = _cell(row, subject_i)
        class_val = _cell(row, class_i)
        if not teacher_val or not subject_val or not class_val:
            continue
        try:
            periods = int(_cell(row, periods_i) or 1)
        except ValueError:
            result.errors.append(RowError(row_idx, "Periods Per Week must be a number."))
            continue
        if periods < 1:
            result.errors.append(RowError(row_idx, "Periods Per Week must be at least 1."))
            continue
        try:
            duration = int(_cell(row, duration_i) or 1) if duration_i >= 0 else 1
        except ValueError:
            result.errors.append(RowError(row_idx, "Duration must be a number."))
            continue
        if duration < 1 or duration > 8:
            result.errors.append(RowError(row_idx, "Duration must be between 1 and 8 periods."))
            continue

        trow = db.fetchone(
            "SELECT id FROM teacher WHERE code = ? OR LOWER(TRIM(first_name || ' ' || last_name)) = ? LIMIT 1",
            (teacher_val.strip(), teacher_val.strip().lower()),
        )
        if not trow:
            result.errors.append(RowError(row_idx, f"Teacher not found: '{teacher_val}'."))
            continue
        teacher_id = trow["id"]

        srow = db.fetchone(
            "SELECT id FROM subject WHERE code = ? OR LOWER(name) = ? LIMIT 1",
            (subject_val.strip(), subject_val.strip().lower()),
        )
        if not srow:
            result.errors.append(RowError(row_idx, f"Subject not found: '{subject_val}'."))
            continue
        subject_id = srow["id"]

        crow = db.fetchone(
            "SELECT id FROM school_class WHERE LOWER(TRIM(name)) = ? OR LOWER(TRIM(grade || ' ' || section)) = ? LIMIT 1",
            (class_val.strip().lower(), class_val.strip().lower()),
        )
        if not crow:
            result.errors.append(RowError(row_idx, f"Class not found: '{class_val}'."))
            continue
        class_id = crow["id"]

        if not dry_run:
            try:
                db.execute(
                    "INSERT INTO lesson (teacher_id, subject_id, class_id, group_id, periods_per_week, duration, priority, locked, preferred_room_id, notes) "
                    "VALUES (?,?,?,NULL,?,?,5,0,NULL,'')",
                    (teacher_id, subject_id, class_id, periods, duration),
                )
                result.success_count += 1
            except Exception as e:
                result.errors.append(RowError(row_idx, str(e)))
        else:
            result.success_count += 1
    if not dry_run:
        db.commit()
    wb.close()
    return result
