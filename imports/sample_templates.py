"""Generate sample Excel templates for import."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font


def write_teachers_template(path: str) -> None:
    """Write sample teachers import template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Teachers"
    headers = ["First Name", "Last Name", "Abbreviation", "Title", "Max Periods Per Day", "Max Periods Per Week"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h).font = Font(bold=True)
    ws.cell(2, 1, "Zain")
    ws.cell(2, 2, "Ahmed")
    ws.cell(2, 3, "ZAI")
    ws.cell(2, 4, "Mr.")
    ws.cell(2, 5, 6)
    ws.cell(2, 6, 28)
    wb.save(path)


def write_classes_template(path: str) -> None:
    """Write sample classes import template (Grade, Section, Stream, Name)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Classes"
    headers = ["Grade", "Section", "Stream", "Name"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h).font = Font(bold=True)
    ws.cell(2, 1, "10")
    ws.cell(2, 2, "A")
    ws.cell(2, 3, "")
    ws.cell(2, 4, "Grade 10-A")
    wb.save(path)


def write_subjects_template(path: str) -> None:
    """Write sample subjects import template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Subjects"
    headers = ["Name", "Code", "Category", "Max Per Day"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h).font = Font(bold=True)
    ws.cell(2, 1, "Mathematics")
    ws.cell(2, 2, "MAT")
    ws.cell(2, 3, "Core")
    ws.cell(2, 4, 2)
    wb.save(path)


def write_lessons_template(path: str) -> None:
    """Write sample lesson mapping template (Teacher, Subject, Class, Periods Per Week, Duration)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Lessons"
    headers = ["Teacher", "Subject", "Class", "Periods Per Week", "Duration"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h).font = Font(bold=True)
    ws.cell(2, 1, "Zain Ahmed")
    ws.cell(2, 2, "Mathematics")
    ws.cell(2, 3, "Grade 10-A")
    ws.cell(2, 4, 5)
    ws.cell(2, 5, 1)
    wb.save(path)
