"""Tests for Excel import and subject library."""
import unittest
import sys
import os
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils.helpers import DEFAULT_SUBJECTS
from imports.excel_import import (
    import_teachers_from_excel,
    import_subjects_from_excel,
    import_lessons_from_excel,
)
from database.connection import DatabaseConnection


class TestDefaultSubjects(unittest.TestCase):
    def test_default_subjects_has_required_entries(self):
        codes = {s[1] for s in DEFAULT_SUBJECTS}
        self.assertIn("MAT", codes)
        self.assertIn("Eng", codes)  # English abbreviation per spec
        self.assertIn("PST", codes)
        self.assertIn("IST", codes)
        self.assertIn("PE", codes)
        self.assertIn("UR", codes)

    def test_default_subjects_structure(self):
        for item in DEFAULT_SUBJECTS:
            self.assertEqual(len(item), 4)
            name, code, category, color = item
            self.assertTrue(len(name) > 0)
            self.assertTrue(len(code) > 0)
            self.assertTrue(color.startswith("#"))

    def test_default_subjects_includes_new_library_entries(self):
        codes = {s[1] for s in DEFAULT_SUBJECTS}
        for code in ("Arts", "Bus", "Com", "Acc", "Add Math", "Hist", "Geo"):
            self.assertIn(code, codes, f"Default subjects should include {code}")


class TestExcelImport(unittest.TestCase):
    def _make_db(self):
        db = DatabaseConnection(":memory:")
        db.open()
        db.initialize_schema()
        db.commit()
        return db

    def test_import_teachers_valid_file(self):
        db = self._make_db()
        db.execute(
            "INSERT INTO school (name, academic_year, days_per_week, periods_per_day, weekend_days) "
            "VALUES (?,?,?,?,?)",
            ("Test", "2025", 5, 7, "5,6"),
        )
        db.commit()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(["First Name", "Last Name", "Abbreviation", "Title", "Max Periods Per Day", "Max Periods Per Week"])
            ws.append(["Ali", "Khan", "AK", "Mr.", 6, 30])
            wb.save(path)
            result = import_teachers_from_excel(db, path)
            self.assertEqual(result.success_count, 1)
            self.assertEqual(len(result.errors), 0)
        finally:
            os.unlink(path)
        db.close()

    def test_import_teachers_missing_first_name_reports_row_error(self):
        db = self._make_db()
        db.execute(
            "INSERT INTO school (name, academic_year, days_per_week, periods_per_day, weekend_days) "
            "VALUES (?,?,?,?,?)",
            ("Test", "2025", 5, 7, "5,6"),
        )
        db.commit()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(["First Name", "Last Name", "Code"])
            ws.append(["", "Khan", "AK"])
            wb.save(path)
            result = import_teachers_from_excel(db, path)
            self.assertEqual(result.success_count, 0)
            self.assertTrue(any(e.row == 2 and "First Name" in e.message for e in result.errors))
        finally:
            os.unlink(path)
        db.close()

    def test_import_subjects_valid(self):
        db = self._make_db()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(["Name", "Code", "Category"])
            ws.append(["Mathematics", "MAT", "Core"])
            wb.save(path)
            result = import_subjects_from_excel(db, path)
            self.assertEqual(result.success_count, 1)
        finally:
            os.unlink(path)
        db.close()

    def test_import_lessons_from_excel_resolves_teacher_subject_class(self):
        db = self._make_db()
        db.execute(
            "INSERT INTO school (name, academic_year, days_per_week, periods_per_day, weekend_days, bell_schedule_json) "
            "VALUES (?,?,?,?,?,?)",
            ("Test", "2025", 5, 7, "5,6", "{}"),
        )
        db.execute(
            "INSERT INTO teacher (first_name, last_name, code, title, max_periods_day, max_periods_week) VALUES (?,?,?,?,?,?)",
            ("Zain", "Ahmed", "ZAI", "Mr.", 6, 28),
        )
        db.execute(
            "INSERT INTO subject (name, code, category, max_per_day) VALUES (?,?,?,?)",
            ("Mathematics", "MAT", "Core", 2),
        )
        db.execute(
            "INSERT INTO school_class (grade, section, name, code, strength) VALUES (?,?,?,?,?)",
            ("10", "A", "Grade 10-A", "10A", 30),
        )
        db.commit()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(["Teacher", "Subject", "Class", "Periods Per Week"])
            ws.append(["Zain Ahmed", "Mathematics", "Grade 10-A", 5])
            wb.save(path)
            result = import_lessons_from_excel(db, path)
            self.assertEqual(result.success_count, 1)
            self.assertEqual(len(result.errors), 0)
            row = db.fetchone("SELECT teacher_id, subject_id, class_id, periods_per_week FROM lesson LIMIT 1")
            self.assertIsNotNone(row)
            self.assertEqual(row["periods_per_week"], 5)
        finally:
            os.unlink(path)
        db.close()

    def test_import_lessons_with_duration_column(self):
        """When Excel has Duration column, it is stored in lesson.duration."""
        db = self._make_db()
        db.execute(
            "INSERT INTO school (name, academic_year, days_per_week, periods_per_day, weekend_days, bell_schedule_json) "
            "VALUES (?,?,?,?,?,?)",
            ("Test", "2025", 5, 7, "5,6", "{}"),
        )
        db.execute(
            "INSERT INTO teacher (first_name, last_name, code, title, max_periods_day, max_periods_week) VALUES (?,?,?,?,?,?)",
            ("Zain", "Ahmed", "ZAI", "Mr.", 6, 28),
        )
        db.execute(
            "INSERT INTO subject (name, code, category, max_per_day) VALUES (?,?,?,?)",
            ("Mathematics", "MAT", "Core", 2),
        )
        db.execute(
            "INSERT INTO school_class (grade, section, stream, name, code, strength) VALUES (?,?,?,?,?,?)",
            ("10", "A", "", "Grade 10-A", "10A", 30),
        )
        db.commit()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(["Teacher", "Subject", "Class", "Periods Per Week", "Duration"])
            ws.append(["Zain Ahmed", "Mathematics", "Grade 10-A", 5, 2])
            wb.save(path)
            result = import_lessons_from_excel(db, path)
            self.assertEqual(result.success_count, 1)
            row = db.fetchone("SELECT periods_per_week, duration FROM lesson LIMIT 1")
            self.assertIsNotNone(row)
            self.assertEqual(row["periods_per_week"], 5)
            self.assertEqual(row["duration"], 2)
        finally:
            os.unlink(path)
        db.close()

    def test_import_teachers_dry_run_does_not_commit(self):
        """dry_run=True returns same validation counts but does not write to DB."""
        db = self._make_db()
        db.execute(
            "INSERT INTO school (name, academic_year, days_per_week, periods_per_day, weekend_days) "
            "VALUES (?,?,?,?,?)",
            ("Test", "2025", 5, 7, "5,6"),
        )
        db.commit()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(["First Name", "Last Name", "Code", "Max Periods Per Day", "Max Periods Per Week"])
            ws.append(["Alice", "Smith", "AS", 6, 28])
            ws.append(["Bob", "Jones", "BJ", 6, 30])
            wb.save(path)
            preview = import_teachers_from_excel(db, path, dry_run=True)
            self.assertEqual(preview.success_count, 2)
            self.assertEqual(preview.total_rows, 2)
            self.assertEqual(len(preview.errors), 0)
            count = db.fetchone("SELECT COUNT(*) as n FROM teacher")
            self.assertEqual(count["n"], 0, "dry_run must not insert any row")
            result = import_teachers_from_excel(db, path, dry_run=False)
            self.assertEqual(result.success_count, 2)
            count = db.fetchone("SELECT COUNT(*) as n FROM teacher")
            self.assertEqual(count["n"], 2)
        finally:
            os.unlink(path)
        db.close()


if __name__ == "__main__":
    unittest.main()
