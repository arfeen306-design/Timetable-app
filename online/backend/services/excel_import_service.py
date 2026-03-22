"""Excel import for teachers and classes — project-scoped, batch-optimized.

Phase 1: Subject column parsing + linking
Phase 2: Batch operations for <2s processing of 500+ rows
"""
from __future__ import annotations
from dataclasses import dataclass, field
from io import BytesIO
from typing import List

from openpyxl import load_workbook
from sqlalchemy.orm import Session
from sqlalchemy import func

from backend.models.project import Subject
from backend.repositories import teacher_repo, class_repo


@dataclass
class RowError:
    row: int  # 1-based Excel row
    message: str


@dataclass
class ImportResult:
    success_count: int = 0
    errors: List[RowError] = field(default_factory=list)
    total_rows: int = 0
    subjects_linked: int = 0

    @property
    def is_valid(self) -> bool:
        return len(self.errors) == 0


def _cell(row: list, col_index: int) -> str:
    if col_index < len(row):
        v = row[col_index]
        return str(v).strip() if v is not None else ""
    return ""


def import_teachers_from_excel(
    db: Session, project_id: int, content: bytes
) -> ImportResult:
    """Parse teachers Excel and create via batch insert with subject linking."""
    result = ImportResult()
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        result.errors.append(RowError(1, "File must have a header row and at least one data row."))
        return result

    data_rows = rows[1:]
    result.total_rows = len(data_rows)
    header = [str(c).strip().lower() if c else "" for c in rows[0]]

    # ── Detect column indices from header ──
    first_i = next((i for i, h in enumerate(header) if "first" in h and "name" in h), 0)
    last_i = next((i for i, h in enumerate(header) if "last" in h and "name" in h), 1)
    code_i = next((i for i, h in enumerate(header) if "abbreviation" in h or "code" in h), 2)
    title_i = next((i for i, h in enumerate(header) if "title" in h), 3)
    max_day_i = next((i for i, h in enumerate(header) if "max" in h and "day" in h), -1)
    max_week_i = next((i for i, h in enumerate(header) if "max" in h and "week" in h), -1)
    subject_i = next((i for i, h in enumerate(header) if "subject" in h), -1)

    # ── Phase 1: Parse all rows into validated records ──
    parsed_rows: list[dict] = []
    for row_idx, row in enumerate(rows[1:], start=2):
        row = list(row) if row else []
        first = _cell(row, first_i)
        last = _cell(row, last_i)
        if not first and not last:
            continue  # skip blank rows
        if not first:
            result.errors.append(RowError(row_idx, "First Name is required."))
            continue
        code = _cell(row, code_i) or (first[:3].upper() if first else "")
        title = _cell(row, title_i) or "Mr."
        try:
            max_day = int(_cell(row, max_day_i) or 6)
            max_week = int(_cell(row, max_week_i) or 30)
        except ValueError:
            result.errors.append(RowError(row_idx, "Max Periods Day/Week must be numbers."))
            continue
        if max_day < 1 or max_week < 1:
            result.errors.append(RowError(row_idx, "Max periods must be at least 1."))
            continue

        subject_name = _cell(row, subject_i) if subject_i >= 0 else ""

        parsed_rows.append({
            "row_idx": row_idx,
            "first": first,
            "last": last,
            "code": code,
            "title": title,
            "max_day": max_day,
            "max_week": max_week,
            "subject_name": subject_name,
        })

    if not parsed_rows:
        return result

    # ── Phase 2a: Bulk-check for duplicate names ──
    existing_names: set[tuple[str, str]] = set()
    existing = teacher_repo.list_by_project(db, project_id)
    for t in existing:
        existing_names.add((t.first_name.strip().lower(), t.last_name.strip().lower()))

    # ── Phase 2b: Bulk-resolve subject names → IDs in one query ──
    unique_subject_names = {r["subject_name"].lower() for r in parsed_rows if r["subject_name"]}
    subject_map: dict[str, int] = {}  # lowercase name → subject.id
    if unique_subject_names:
        subjects = (
            db.query(Subject)
            .filter(Subject.project_id == project_id, func.lower(Subject.name).in_(unique_subject_names))
            .all()
        )
        for s in subjects:
            subject_map[s.name.strip().lower()] = s.id

    # ── Phase 2c: Batch insert teachers ──
    created_teachers: list[tuple[int, str]] = []  # (teacher_id, subject_name)
    for r in parsed_rows:
        key = (r["first"].strip().lower(), r["last"].strip().lower())
        if key in existing_names:
            result.errors.append(RowError(r["row_idx"], "A teacher with this first and last name already exists."))
            continue

        # Validate subject exists if provided
        subj_name = r["subject_name"]
        if subj_name and subj_name.lower() not in subject_map:
            result.errors.append(RowError(
                r["row_idx"],
                f'Subject "{subj_name}" not found in project. Add it first under Subjects tab.'
            ))
            continue

        try:
            teacher = teacher_repo.create(
                db,
                project_id=project_id,
                first_name=r["first"],
                last_name=r["last"],
                code=r["code"],
                title=r["title"],
                max_periods_day=r["max_day"],
                max_periods_week=r["max_week"],
            )
            existing_names.add(key)
            result.success_count += 1
            if subj_name:
                created_teachers.append((teacher.id, subj_name))
        except Exception as e:
            result.errors.append(RowError(r["row_idx"], str(e)))

    # ── Phase 2d: Bulk-link teacher ↔ subject ──
    for teacher_id, subj_name in created_teachers:
        sid = subject_map.get(subj_name.lower())
        if sid:
            try:
                teacher_repo.set_teacher_subjects(db, teacher_id, [sid])
                result.subjects_linked += 1
            except Exception:
                pass  # non-critical — teacher already created

    return result


def import_classes_from_excel(
    db: Session, project_id: int, content: bytes
) -> ImportResult:
    """Parse classes Excel (desktop template format) and create via class_repo."""
    result = ImportResult()
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        result.errors.append(RowError(1, "File must have a header row and at least one data row."))
        return result

    result.total_rows = len(rows) - 1
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    grade_i = next((i for i, h in enumerate(header) if "grade" in h), 0)
    section_i = next((i for i, h in enumerate(header) if "section" in h), 1)
    stream_i = next((i for i, h in enumerate(header) if "stream" in h), -1)
    name_i = next((i for i, h in enumerate(header) if "name" in h), 2)
    if name_i >= len(header):
        name_i = max(grade_i, section_i, stream_i if stream_i >= 0 else 0) + 1

    for row_idx, row in enumerate(rows[1:], start=2):
        row = list(row) if row else []
        grade = _cell(row, grade_i)
        section = _cell(row, section_i)
        stream = _cell(row, stream_i) if stream_i >= 0 else ""
        name = _cell(row, name_i)
        if not name:
            name = f"Grade {grade} {section}".strip() or f"Row {row_idx}"
        if class_repo.find_by_grade_section_stream(db, project_id, grade, section, stream):
            result.errors.append(
                RowError(row_idx, "A class with this grade, section and stream already exists.")
            )
            continue
        try:
            code = (str(grade or "") + str(section or "")).replace(" ", "")[:10] or name[:6]
            class_repo.create(
                db,
                project_id=project_id,
                grade=grade or "",
                section=section or "",
                stream=stream or "",
                name=name,
                code=code,
                strength=30,
            )
            result.success_count += 1
        except Exception as e:
            result.errors.append(RowError(row_idx, str(e)))
    return result
