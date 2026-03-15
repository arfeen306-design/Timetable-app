"""Direct exports — Excel, CSV, PDF from Postgres. Includes time slots + breaks."""
from __future__ import annotations
import io, json
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from backend.auth.project_scope import get_project_or_404
from backend.models.base import get_db
from backend.models.project import Project
from backend.repositories.timetable_repo import get_latest_run, get_entries_with_joins
from backend.repositories.school_settings_repo import get_by_project

router = APIRouter()

DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
DAY_SHORT = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


# ═══════════════════════════════════════════════════════════════════════════════
# Slot computation — matches frontend computeSlots() exactly
# ═══════════════════════════════════════════════════════════════════════════════

def _parse_time(t: str) -> int:
    if not t or ":" not in t:
        return 8 * 60
    parts = t.split(":")
    return int(parts[0]) * 60 + int(parts[1])

def _fmt_time(mins: int) -> str:
    return f"{mins // 60:02d}:{mins % 60:02d}"

def _compute_col_headers(settings) -> list[dict]:
    """
    Returns list of column dicts: {type, label, sub, period_index}
    type = "period" | "break" | "zero"
    """
    school_start = getattr(settings, "school_start_time", None) or "08:00"
    period_dur = getattr(settings, "period_duration_minutes", None) or 45
    num_periods = getattr(settings, "periods_per_day", None) or 7

    breaks = []
    try:
        raw = getattr(settings, "breaks_json", None) or "[]"
        breaks = json.loads(raw)
    except: pass

    bell = {}
    try:
        raw = getattr(settings, "bell_schedule_json", None) or "{}"
        bell = json.loads(raw)
    except: pass

    first_period = bell.get("first_period_start") or school_start
    zero_period = bool(bell.get("zero_period"))

    # Regular breaks (not friday)
    day_breaks = [b for b in breaks if not b.get("is_friday")]
    break_by_idx = {}
    for b in day_breaks:
        ap = b.get("after_period")
        if ap is not None:
            break_by_idx[ap - 1] = b  # convert 1-indexed to 0-indexed

    cols = []
    # Zero period
    if zero_period:
        ss = _parse_time(school_start)
        fp = _parse_time(first_period)
        if ss < fp:
            cols.append({"type": "zero", "label": "0", "sub": f"{_fmt_time(ss)} TO {_fmt_time(fp)}", "period_index": -1})

    current = _parse_time(first_period)
    for p in range(num_periods):
        end = current + period_dur
        cols.append({"type": "period", "label": str(p + 1), "sub": f"{_fmt_time(current)} TO {_fmt_time(end)}", "period_index": p})
        current = end
        brk = break_by_idx.get(p)
        if brk:
            bs = _parse_time(brk.get("start", "")) if brk.get("start") else current
            be = _parse_time(brk.get("end", "")) if brk.get("end") else bs + (brk.get("duration_minutes") or 20)
            name = brk.get("name") or "Break"
            cols.append({"type": "break", "label": name.upper(), "sub": f"{_fmt_time(bs)} TO {_fmt_time(be)}", "period_index": -1})
            current = be

    return cols

def _compute_friday_headers(settings) -> list[dict] | None:
    """Returns Friday-specific column headers, or None if no Friday difference."""
    bell = {}
    try:
        raw = getattr(settings, "bell_schedule_json", None) or "{}"
        bell = json.loads(raw)
    except: pass

    if not bell.get("friday_different"):
        return None

    school_start = getattr(settings, "school_start_time", None) or "08:00"
    period_dur = getattr(settings, "period_duration_minutes", None) or 45
    num_periods = getattr(settings, "periods_per_day", None) or 7

    fri_start = bell.get("friday_first_period_start") or bell.get("first_period_start") or school_start
    fri_dur = bell.get("friday_period_duration") or period_dur
    zero_period = bool(bell.get("zero_period"))

    breaks = []
    try:
        raw = getattr(settings, "breaks_json", None) or "[]"
        breaks = json.loads(raw)
    except: pass

    fri_breaks = [b for b in breaks if b.get("is_friday")]
    break_by_idx = {}
    for b in fri_breaks:
        ap = b.get("after_period")
        if ap is not None:
            break_by_idx[ap - 1] = b

    cols = []
    if zero_period:
        ss = _parse_time(school_start)
        fp = _parse_time(fri_start)
        if ss < fp:
            cols.append({"type": "zero", "label": "0", "sub": f"{_fmt_time(ss)} TO {_fmt_time(fp)}", "period_index": -1})

    current = _parse_time(fri_start)
    for p in range(num_periods):
        end = current + fri_dur
        cols.append({"type": "period", "label": str(p + 1), "sub": f"{_fmt_time(current)} TO {_fmt_time(end)}", "period_index": p})
        current = end
        brk = break_by_idx.get(p)
        if brk:
            bs = _parse_time(brk.get("start", "")) if brk.get("start") else current
            be = _parse_time(brk.get("end", "")) if brk.get("end") else bs + (brk.get("duration_minutes") or 20)
            name = brk.get("name") or "Break"
            cols.append({"type": "break", "label": name.upper(), "sub": f"{_fmt_time(bs)} TO {_fmt_time(be)}", "period_index": -1})
            current = be

    return cols


def _ensure_completed_run(db: Session, project_id: int):
    run = get_latest_run(db, project_id)
    if not run or run.status != "completed":
        raise HTTPException(400, "No completed timetable run. Generate first.")
    return run


def _get_all_entries(db: Session, project_id: int, run_id: int) -> list[dict]:
    return get_entries_with_joins(db, project_id, run_id=run_id)


def _full_settings(db: Session, project_id: int):
    return get_by_project(db, project_id)


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL — Sheet1: All Classes, Sheet2: All Teachers (with time slots + breaks)
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/excel")
def export_excel(
    project: Project = Depends(get_project_or_404),
    db: Session = Depends(get_db),
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(501, "openpyxl not installed on server.")

    run = _ensure_completed_run(db, project.id)
    entries = _get_all_entries(db, project.id, run.id)
    sett = _full_settings(db, project.id)
    days = getattr(sett, "days_per_week", 5)
    school_name = getattr(sett, "name", "") or ""

    cols = _compute_col_headers(sett)
    fri_cols = _compute_friday_headers(sett)
    bell = {}
    try: bell = json.loads(getattr(sett, "bell_schedule_json", "{}") or "{}")
    except: pass
    fri_day_idx = bell.get("friday_day_index", 4)

    by_class: dict[str, list[dict]] = {}
    by_teacher: dict[str, list[dict]] = {}
    for e in entries:
        by_class.setdefault(e.get("class_name", "?"), []).append(e)
        by_teacher.setdefault(e.get("teacher_name", "?"), []).append(e)

    wb = Workbook()
    hdr_font = Font(bold=True, size=9, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2C3E50")
    brk_fill = PatternFill("solid", fgColor="4A5568")
    zero_fill = PatternFill("solid", fgColor="2C5282")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    gray_fill = PatternFill("solid", fgColor="E2E8F0")

    def _write_sheet(ws, title: str, groups: dict[str, list[dict]], label_key: str):
        """Write all entities vertically, time slots horizontally."""
        ws.cell(1, 1, f"{school_name} — {title}").font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols) + 2)
        row = 3

        for entity_name in sorted(groups.keys()):
            group_entries = groups[entity_name]
            lookup = {(e["day_index"], e["period_index"]): e for e in group_entries}

            # Entity name header
            ws.cell(row, 1, entity_name).font = Font(bold=True, size=11, color="1E293B")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(cols) + 1)
            row += 1

            # Column headers: Day + slots
            ws.cell(row, 1, "Day").font = hdr_font
            ws.cell(row, 1).fill = hdr_fill
            ws.cell(row, 1).alignment = hdr_align
            ws.cell(row, 1).border = thin
            ws.column_dimensions["A"].width = 8
            for ci, col in enumerate(cols):
                c = ws.cell(row, ci + 2)
                c.value = f"{col['label']}\n{col['sub']}"
                c.font = hdr_font
                c.fill = brk_fill if col["type"] == "break" else zero_fill if col["type"] == "zero" else hdr_fill
                c.alignment = hdr_align
                c.border = thin
                ws.column_dimensions[get_column_letter(ci + 2)].width = 16
            row += 1

            # Day rows
            for d in range(days):
                is_fri = d == fri_day_idx
                ws.cell(row, 1, DAY_SHORT[d]).font = Font(bold=True)
                ws.cell(row, 1).alignment = cell_align
                ws.cell(row, 1).border = thin
                for ci, col in enumerate(cols):
                    cell = ws.cell(row, ci + 2)
                    cell.border = thin
                    cell.alignment = cell_align
                    if col["type"] == "break":
                        cell.fill = gray_fill
                        cell.value = col["label"].lower()
                        cell.font = Font(italic=True, size=8, color="64748B")
                    elif col["type"] == "zero":
                        cell.fill = PatternFill("solid", fgColor="EBF8FF")
                        cell.font = Font(italic=True, size=8, color="2B6CB0")
                    else:
                        e = lookup.get((d, col["period_index"]))
                        if e:
                            code = e.get("subject_code") or e.get("subject_name", "")
                            lbl = e.get(label_key, "")
                            room = e.get("room_name", "")
                            text = code
                            if lbl: text += f"\n{lbl}"
                            if room: text += f"\n{room}"
                            cell.value = text
                            color = (e.get("subject_color") or "#4A90D9").lstrip("#")
                            try:
                                ri, gi, bi = int(color[:2],16), int(color[2:4],16), int(color[4:6],16)
                                lr, lg, lb = min(ri+60,255), min(gi+60,255), min(bi+60,255)
                                cell.fill = PatternFill("solid", fgColor=f"{lr:02X}{lg:02X}{lb:02X}")
                            except: pass
                row += 1

            # Friday times row if different
            if fri_cols:
                ws.cell(row, 1, "Fri times").font = Font(bold=True, size=8, color="92400E")
                ws.cell(row, 1).fill = PatternFill("solid", fgColor="FEF3C7")
                ws.cell(row, 1).border = thin
                for ci, fc in enumerate(fri_cols):
                    if ci < len(cols):
                        cell = ws.cell(row, ci + 2)
                        cell.value = fc["sub"]
                        cell.font = Font(size=7, color="92400E")
                        cell.fill = PatternFill("solid", fgColor="FEF3C7")
                        cell.alignment = cell_align
                        cell.border = thin
                row += 1

            row += 1  # gap between entities

    # Sheet 1: All Classes
    ws_classes = wb.active
    ws_classes.title = "All Classes"
    _write_sheet(ws_classes, "All Class Timetables", by_class, "teacher_name")

    # Sheet 2: All Teachers
    ws_teachers = wb.create_sheet(title="All Teachers")
    _write_sheet(ws_teachers, "All Teacher Timetables", by_teacher, "class_name")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"timetable_{project.name or project.id}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ═══════════════════════════════════════════════════════════════════════════════
# CSV — All Classes + All Teachers flat (with time slots)
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/csv")
def export_csv(
    project: Project = Depends(get_project_or_404),
    db: Session = Depends(get_db),
):
    import csv as csv_mod

    run = _ensure_completed_run(db, project.id)
    entries = _get_all_entries(db, project.id, run.id)
    sett = _full_settings(db, project.id)
    cols = _compute_col_headers(sett)
    days = getattr(sett, "days_per_week", 5)

    # Build period_index → time slot lookup
    slot_times = {}
    for col in cols:
        if col["type"] == "period":
            slot_times[col["period_index"]] = col["sub"]

    buf = io.StringIO()
    writer = csv_mod.writer(buf)
    writer.writerow(["Type", "Name", "Day", "Period", "Time Slot", "Subject", "Subject Code", "Teacher", "Class", "Room"])

    # Classes
    by_class: dict[str, list[dict]] = {}
    by_teacher: dict[str, list[dict]] = {}
    for e in entries:
        by_class.setdefault(e.get("class_name", "?"), []).append(e)
        by_teacher.setdefault(e.get("teacher_name", "?"), []).append(e)

    for cn in sorted(by_class.keys()):
        for e in sorted(by_class[cn], key=lambda x: (x["day_index"], x["period_index"])):
            writer.writerow([
                "Class", cn,
                DAY_SHORT[e["day_index"]] if e["day_index"] < len(DAY_SHORT) else f"Day{e['day_index']+1}",
                e["period_index"] + 1,
                slot_times.get(e["period_index"], ""),
                e.get("subject_name", ""), e.get("subject_code", ""),
                e.get("teacher_name", ""), e.get("class_name", ""), e.get("room_name", ""),
            ])

    for tn in sorted(by_teacher.keys()):
        for e in sorted(by_teacher[tn], key=lambda x: (x["day_index"], x["period_index"])):
            writer.writerow([
                "Teacher", tn,
                DAY_SHORT[e["day_index"]] if e["day_index"] < len(DAY_SHORT) else f"Day{e['day_index']+1}",
                e["period_index"] + 1,
                slot_times.get(e["period_index"], ""),
                e.get("subject_name", ""), e.get("subject_code", ""),
                e.get("teacher_name", ""), e.get("class_name", ""), e.get("room_name", ""),
            ])

    content = buf.getvalue().encode("utf-8")
    fname = f"timetable_{project.name or project.id}.csv"
    return StreamingResponse(io.BytesIO(content),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ═══════════════════════════════════════════════════════════════════════════════
# PDF helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _build_pdf_table(entries_group, cols, days, fri_day_idx, fri_cols, label_key):
    """Build a table data array for one entity."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm

    lookup = {(e["day_index"], e["period_index"]): e for e in entries_group}
    header = ["Day"] + [f"{c['label']}\n{c['sub']}" for c in cols]
    data = [header]

    for d in range(days):
        row = [DAY_SHORT[d]]
        for c in cols:
            if c["type"] == "break":
                row.append(c["label"].lower())
            elif c["type"] == "zero":
                row.append("")
            else:
                e = lookup.get((d, c["period_index"]))
                if e:
                    code = e.get("subject_code") or e.get("subject_name", "")
                    lbl = e.get(label_key, "")
                    room = e.get("room_name", "")
                    t = code
                    if lbl: t += f"\n{lbl}"
                    if room: t += f"\n{room}"
                    row.append(t)
                else:
                    row.append("")
        data.append(row)

    # Friday times row
    if fri_cols and fri_day_idx < days:
        fri_row = ["Fri times"]
        for i, c in enumerate(cols):
            if i < len(fri_cols):
                fri_row.append(fri_cols[i]["sub"])
            else:
                fri_row.append("")
        data.insert(fri_day_idx + 1, fri_row)  # insert before Friday data row

    return data


def _style_pdf_table(tbl, cols, has_fri_row, fri_row_idx):
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle

    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("FONTSIZE", (0, 1), (-1, -1), 6.5),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#F1F5F9")),
    ]

    # Break column backgrounds in header
    for ci, c in enumerate(cols):
        if c["type"] == "break":
            style_cmds.append(("BACKGROUND", (ci + 1, 0), (ci + 1, 0), colors.HexColor("#4A5568")))
            # Grey for all rows in break column
            style_cmds.append(("BACKGROUND", (ci + 1, 1), (ci + 1, -1), colors.HexColor("#E2E8F0")))
        elif c["type"] == "zero":
            style_cmds.append(("BACKGROUND", (ci + 1, 0), (ci + 1, 0), colors.HexColor("#2C5282")))
            style_cmds.append(("BACKGROUND", (ci + 1, 1), (ci + 1, -1), colors.HexColor("#EBF8FF")))

    # Friday times row styling
    if has_fri_row:
        r = fri_row_idx + 1  # +1 for header
        style_cmds.append(("BACKGROUND", (0, r), (-1, r), colors.HexColor("#FEF3C7")))
        style_cmds.append(("TEXTCOLOR", (0, r), (-1, r), colors.HexColor("#92400E")))
        style_cmds.append(("FONTSIZE", (0, r), (-1, r), 5.5))

    tbl.setStyle(TableStyle(style_cmds))


# ═══════════════════════════════════════════════════════════════════════════════
# PDF — Current view only (class_id, teacher_id, or room_id)
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/pdf")
def export_pdf_current(
    project: Project = Depends(get_project_or_404),
    db: Session = Depends(get_db),
    class_id: int = Query(0),
    teacher_id: int = Query(0),
    room_id: int = Query(0),
):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        raise HTTPException(501, "reportlab not installed on server.")

    run = _ensure_completed_run(db, project.id)
    entries = _get_all_entries(db, project.id, run.id)
    sett = _full_settings(db, project.id)
    days = getattr(sett, "days_per_week", 5)
    school_name = getattr(sett, "name", "") or ""
    cols = _compute_col_headers(sett)
    fri_cols = _compute_friday_headers(sett)
    bell = {}
    try: bell = json.loads(getattr(sett, "bell_schedule_json", "{}") or "{}")
    except: pass
    fri_day_idx = bell.get("friday_day_index", 4)

    # Filter entries
    if class_id:
        filtered = [e for e in entries if e.get("class_id") == class_id]
        label_key = "teacher_name"
        entity_name = filtered[0]["class_name"] if filtered else f"Class {class_id}"
        title = f"Class: {entity_name}"
    elif teacher_id:
        filtered = [e for e in entries if e.get("teacher_id") == teacher_id]
        label_key = "class_name"
        entity_name = filtered[0]["teacher_name"] if filtered else f"Teacher {teacher_id}"
        title = f"Teacher: {entity_name}"
    elif room_id:
        filtered = [e for e in entries if e.get("room_id") == room_id]
        label_key = "class_name"
        entity_name = filtered[0].get("room_name", f"Room {room_id}") if filtered else f"Room {room_id}"
        title = f"Room: {entity_name}"
    else:
        # Default: all classes
        filtered = entries
        label_key = "teacher_name"
        title = "All Classes"

    data = _build_pdf_table(filtered, cols, days, fri_day_idx, fri_cols, label_key)
    has_fri = fri_cols is not None and fri_day_idx < days

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=12*mm, bottomMargin=12*mm, leftMargin=8*mm, rightMargin=8*mm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(school_name, ParagraphStyle("T", parent=styles["Title"], fontSize=16, spaceAfter=2)))
    elements.append(Paragraph(title, ParagraphStyle("S", parent=styles["Heading2"], fontSize=13, spaceAfter=4)))
    elements.append(Spacer(1, 3))

    page_w = landscape(A4)[0] - 16*mm
    day_w = 22*mm
    col_w = (page_w - day_w) / len(cols)
    tbl = Table(data, colWidths=[day_w] + [col_w] * len(cols), repeatRows=1)
    _style_pdf_table(tbl, cols, has_fri, fri_day_idx)
    elements.append(tbl)

    doc.build(elements)
    buf.seek(0)
    fname = f"timetable_view_{project.name or project.id}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ═══════════════════════════════════════════════════════════════════════════════
# PDF — All Classes (one per page)
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/pdf/all-classes")
def export_pdf_all_classes(
    project: Project = Depends(get_project_or_404),
    db: Session = Depends(get_db),
):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        raise HTTPException(501, "reportlab not installed on server.")

    run = _ensure_completed_run(db, project.id)
    entries = _get_all_entries(db, project.id, run.id)
    sett = _full_settings(db, project.id)
    days = getattr(sett, "days_per_week", 5)
    school_name = getattr(sett, "name", "") or ""
    cols = _compute_col_headers(sett)
    fri_cols = _compute_friday_headers(sett)
    bell = {}
    try: bell = json.loads(getattr(sett, "bell_schedule_json", "{}") or "{}")
    except: pass
    fri_day_idx = bell.get("friday_day_index", 4)

    by_class: dict[str, list[dict]] = {}
    for e in entries:
        by_class.setdefault(e.get("class_name", "?"), []).append(e)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=12*mm, bottomMargin=12*mm, leftMargin=8*mm, rightMargin=8*mm)
    styles = getSampleStyleSheet()
    elements = []
    title_s = ParagraphStyle("T", parent=styles["Title"], fontSize=16, spaceAfter=2)
    sub_s = ParagraphStyle("S", parent=styles["Heading2"], fontSize=13, spaceAfter=4)
    page_w = landscape(A4)[0] - 16*mm
    day_w = 22*mm
    col_w = (page_w - day_w) / len(cols)
    has_fri = fri_cols is not None and fri_day_idx < days

    for idx, cn in enumerate(sorted(by_class.keys())):
        elements.append(Paragraph(school_name, title_s))
        elements.append(Paragraph(f"Class: {cn}", sub_s))
        elements.append(Spacer(1, 3))
        data = _build_pdf_table(by_class[cn], cols, days, fri_day_idx, fri_cols, "teacher_name")
        tbl = Table(data, colWidths=[day_w] + [col_w] * len(cols), repeatRows=1)
        _style_pdf_table(tbl, cols, has_fri, fri_day_idx)
        elements.append(tbl)
        if idx < len(by_class) - 1:
            elements.append(PageBreak())

    doc.build(elements)
    buf.seek(0)
    fname = f"all_classes_{project.name or project.id}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ═══════════════════════════════════════════════════════════════════════════════
# PDF — All Teachers (one per page)
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/pdf/all-teachers")
def export_pdf_all_teachers(
    project: Project = Depends(get_project_or_404),
    db: Session = Depends(get_db),
):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        raise HTTPException(501, "reportlab not installed on server.")

    run = _ensure_completed_run(db, project.id)
    entries = _get_all_entries(db, project.id, run.id)
    sett = _full_settings(db, project.id)
    days = getattr(sett, "days_per_week", 5)
    school_name = getattr(sett, "name", "") or ""
    cols = _compute_col_headers(sett)
    fri_cols = _compute_friday_headers(sett)
    bell = {}
    try: bell = json.loads(getattr(sett, "bell_schedule_json", "{}") or "{}")
    except: pass
    fri_day_idx = bell.get("friday_day_index", 4)

    by_teacher: dict[str, list[dict]] = {}
    for e in entries:
        by_teacher.setdefault(e.get("teacher_name", "?"), []).append(e)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=12*mm, bottomMargin=12*mm, leftMargin=8*mm, rightMargin=8*mm)
    styles = getSampleStyleSheet()
    elements = []
    title_s = ParagraphStyle("T", parent=styles["Title"], fontSize=16, spaceAfter=2)
    sub_s = ParagraphStyle("S", parent=styles["Heading2"], fontSize=13, spaceAfter=4)
    page_w = landscape(A4)[0] - 16*mm
    day_w = 22*mm
    col_w = (page_w - day_w) / len(cols)
    has_fri = fri_cols is not None and fri_day_idx < days

    for idx, tn in enumerate(sorted(by_teacher.keys())):
        elements.append(Paragraph(school_name, title_s))
        elements.append(Paragraph(f"Teacher: {tn}", sub_s))
        elements.append(Spacer(1, 3))
        data = _build_pdf_table(by_teacher[tn], cols, days, fri_day_idx, fri_cols, "class_name")
        tbl = Table(data, colWidths=[day_w] + [col_w] * len(cols), repeatRows=1)
        _style_pdf_table(tbl, cols, has_fri, fri_day_idx)
        elements.append(tbl)
        if idx < len(by_teacher) - 1:
            elements.append(PageBreak())

    doc.build(elements)
    buf.seek(0)
    fname = f"all_teachers_{project.name or project.id}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.get("")
def list_exports(project: Project = Depends(get_project_or_404)):
    return {"exports": []}
