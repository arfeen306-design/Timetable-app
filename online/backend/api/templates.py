"""Download Excel templates for bulk import — same format as desktop app."""
from __future__ import annotations
from io import BytesIO

from fastapi import APIRouter
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font

router = APIRouter()


def _teachers_workbook() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Teachers"
    headers = ["First Name", "Last Name", "Abbreviation", "Title", "Max Lessons Per Day", "Max Lessons Per Week", "Subject"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h).font = Font(bold=True)
    ws.cell(2, 1, "Zain")
    ws.cell(2, 2, "Ahmed")
    ws.cell(2, 3, "ZAI")
    ws.cell(2, 4, "Mr.")
    ws.cell(2, 5, 6)
    ws.cell(2, 6, 28)
    ws.cell(2, 7, "Mathematics")
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _classes_workbook() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Classes"
    headers = ["Grade", "Section", "Stream", "Name"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h).font = Font(bold=True)
    ws.cell(2, 1, "10")
    ws.cell(2, 2, "A")
    ws.cell(2, 3, "")
    ws.cell(2, 4, "Grade 10-A")
    ws.cell(3, 1, "AS")
    ws.cell(3, 2, "Eng")
    ws.cell(3, 3, "Engineering")
    ws.cell(3, 4, "AS-Eng")
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _rooms_workbook() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Classrooms"
    headers = ["Name", "Code", "Type", "Capacity"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h).font = Font(bold=True)
    ws.cell(2, 1, "Room 101")
    ws.cell(2, 2, "R101")
    ws.cell(2, 3, "Classroom")
    ws.cell(2, 4, 35)
    ws.cell(3, 1, "Physics Lab")
    ws.cell(3, 2, "PLAB")
    ws.cell(3, 3, "Lab")
    ws.cell(3, 4, 24)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/teachers.xlsx")
def download_teachers_template():
    buf = _teachers_workbook()
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=teachers_template.xlsx"},
    )


@router.get("/classes.xlsx")
def download_classes_template():
    buf = _classes_workbook()
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=classes_template.xlsx"},
    )


@router.get("/rooms.xlsx")
def download_rooms_template():
    buf = _rooms_workbook()
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=rooms_template.xlsx"},
    )
